import sys
import customtkinter as ctk
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
import shutil
import threading
import subprocess
import os

# Set CustomTkinter appearance and color theme
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

# Get the directory where the exe is located
if getattr(sys, 'frozen', False):
    BASE_DIR = Path(sys.executable).parent
else:
    BASE_DIR = Path(__file__).parent

INPUT_DIR = BASE_DIR / "inputdir" / "Returnsreportfiles"
TEMPLATE_PATH = BASE_DIR / "Template" / "ReturnsReconcileReport.xlsx"
OUTPUT_DIR = BASE_DIR / "Output"

# GUI Color scheme - Orange and White (matching main app)
COLORS = {
    "primary_orange": "#FF6B35",
    "secondary_orange": "#FF8F65",
    "light_orange": "#FFB399",
    "white": "#FFFFFF",
    "light_gray": "#F5F5F5",
    "dark_gray": "#333333",
    "text_dark": "#2D2D2D",
    "success_green": "#4CAF50",
    "warning_orange": "#FF9800",
    "error_red": "#F44336",
    "accent_blue": "#2196F3"
}

class ReturnsReportGUI:
    def __init__(self, parent_window=None, back_callback=None):
        """
        Initialize the Returns module
        Args:
            parent_window: Reference to the main homepage window
            back_callback: Function to call when back button is pressed
        """
        self.parent_window = parent_window
        self.back_callback = back_callback
        
        # Create window or use parent
        if parent_window is None:
            # Standalone mode
            self.root = ctk.CTk()
            self.root.title("Returns Reconciliation Module")
            self.root.geometry("900x700")
            self.root.minsize(800, 600)
            self.is_standalone = True
        else:
            # Integrated mode - use parent window
            self.root = parent_window
            self.is_standalone = False
        
        # Configure colors
        self.root.configure(fg_color=COLORS["white"])
        
        # File checkboxes dictionary
        self.file_checkboxes = {}
        
        # Progress variables
        self.progress_var = None
        self.status_label = None
        
        # Store original content if integrating
        self.original_content = None
        if not self.is_standalone:
            self.store_original_content()
        
        self.setup_gui()
        self.update_file_status()
        
    def store_original_content(self):
        """Store the original homepage content to restore later"""
        if self.parent_window:
            self.original_content = []
            for child in self.parent_window.winfo_children():
                self.original_content.append(child)
                child.pack_forget()  # Hide original content
                
    def restore_original_content(self):
        """Restore the original homepage content"""
        if self.original_content:
            # Clear current content
            for child in self.root.winfo_children():
                child.destroy()
            
            # Restore original content
            for child in self.original_content:
                child.pack(fill="both", expand=True)
                
    def setup_gui(self):
        """Setup the main GUI layout"""
        # Main container
        main_frame = ctk.CTkFrame(self.root, fg_color=COLORS["white"], corner_radius=0)
        main_frame.pack(fill="both", expand=True, padx=30, pady=30)
        
        # Header section
        self.create_header(main_frame)

        
        # Main content area
        content_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        content_frame.pack(fill="both", expand=True, pady=(20, 0))

        # Create main sections
        self.create_file_status_section(content_frame)
        self.create_processing_section(content_frame)
        self.create_output_section(content_frame)
        
        # Footer
        self.create_footer(main_frame)
        
    def create_header(self, parent):
        """Create the header section"""
        # Main header
        header_frame = ctk.CTkFrame(parent, fg_color=COLORS["primary_orange"], corner_radius=15, height=100)
        header_frame.pack(fill="x", pady=(0, 20))
        header_frame.pack_propagate(False)
        # Back button
        if self.back_callback:
            back_btn = ctk.CTkButton(
                header_frame,
                text="← Back",
                command=self.back_callback,
                width=80,
                height=30,
                font=ctk.CTkFont(size=14, weight="bold"),
                fg_color=COLORS["accent_blue"],
                hover_color=COLORS["light_orange"]
            )
            back_btn.pack(side="left", padx=20, pady=15)
        # Title and subtitle
        title_container = ctk.CTkFrame(header_frame, fg_color="transparent")
        title_container.pack(expand=True, fill="both")
        
        title_label = ctk.CTkLabel(
            title_container, 
            text="↩️ Returns Reconciliation",
            font=ctk.CTkFont(size=28, weight="bold"),
            text_color=COLORS["white"]
        )
        title_label.pack(pady=(15, 5))
        
        subtitle_label = ctk.CTkLabel(
            title_container,
            text="Reconcile return shipments and generate comprehensive reports",
            font=ctk.CTkFont(size=14),
            text_color=COLORS["white"]
        )
        subtitle_label.pack(pady=(0, 15))

        
    def create_file_status_section(self, parent):
        """Create file status checking section"""
        # File Status Frame
        status_frame = ctk.CTkFrame(parent, fg_color=COLORS["light_gray"], corner_radius=12)
        status_frame.pack(fill="x", pady=(0, 15))
        
        # Header with refresh button
        header_frame = ctk.CTkFrame(status_frame, fg_color="transparent")
        header_frame.pack(fill="x", padx=20, pady=(15, 10))
        
        status_title = ctk.CTkLabel(
            header_frame,
            text="📁 Required Files Status",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color=COLORS["text_dark"]
        )
        status_title.pack(side="left")
        
        # Buttons frame for refresh and reset
        buttons_frame = ctk.CTkFrame(header_frame, fg_color="transparent")
        buttons_frame.pack(side="right")
        
        # Refresh button
        refresh_btn = ctk.CTkButton(
            buttons_frame,
            text="🔄",
            command=self.update_file_status,
            width=30,
            height=30,
            font=ctk.CTkFont(size=14),
            fg_color=COLORS["accent_blue"],
            hover_color="#1976D2"
        )
        refresh_btn.pack(side="right", padx=(0, 5))
        
        # Reset Files button
        reset_btn = ctk.CTkButton(
            buttons_frame,
            text="🗑️ Reset Files",
            command=self.reset_files,
            width=100,
            height=30,
            font=ctk.CTkFont(size=11),
            fg_color=COLORS["error_red"],
            hover_color="#D32F2F",
            text_color=COLORS["white"]
        )
        reset_btn.pack(side="right")
        
        # File checkboxes frame
        files_frame = ctk.CTkFrame(status_frame, fg_color="transparent")
        files_frame.pack(fill="x", padx=20, pady=(0, 15))
        
        # Configure grid
        files_frame.grid_columnconfigure(0, weight=1)
        files_frame.grid_columnconfigure(1, weight=1)
        
        # File list for returns
        required_files = [
            "Returns Meesho.csv",
            "Returns Flipkart KC.csv", 
            "Returns Flipkart LL.csv",
            "Returns SellerFlex.csv"
        ]
        
        for i, filename in enumerate(required_files):
            row = i // 2
            col = i % 2
            
            checkbox = ctk.CTkCheckBox(
                files_frame,
                text=filename,
                font=ctk.CTkFont(size=12),
                text_color=COLORS["text_dark"],
                state="disabled"
            )
            checkbox.grid(row=row, column=col, sticky="w", padx=10, pady=5)
            self.file_checkboxes[filename] = checkbox
            
    def create_processing_section(self, parent):
        """Create processing control section"""
        process_frame = ctk.CTkFrame(parent, fg_color=COLORS["light_gray"], corner_radius=12)
        process_frame.pack(fill="x", pady=(0, 15))
        
        # Section title
        process_title = ctk.CTkLabel(
            process_frame,
            text="⚙️ Processing Controls",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color=COLORS["text_dark"]
        )
        process_title.pack(pady=(15, 10))
        
        # Description
        desc_label = ctk.CTkLabel(
            process_frame,
            text="Process return data from Meesho, Flipkart KC/LL, and SellerFlex to generate consolidated reports",
            font=ctk.CTkFont(size=12),
            text_color=COLORS["dark_gray"],
            wraplength=600
        )
        desc_label.pack(pady=(0, 15))
        
        # Process button
        self.process_btn = ctk.CTkButton(
            process_frame,
            text="🚀 Start Returns Processing",
            command=self.start_processing,
            width=250,
            height=45,
            font=ctk.CTkFont(size=14, weight="bold"),
            fg_color=COLORS["primary_orange"],
            hover_color=COLORS["secondary_orange"]
        )
        self.process_btn.pack(pady=(0, 20))
        
    def create_output_section(self, parent):
        """Create output and progress section"""
        output_frame = ctk.CTkFrame(parent, fg_color=COLORS["light_gray"], corner_radius=12)
        output_frame.pack(fill="both", expand=True)
        
        # Section title
        output_title = ctk.CTkLabel(
            output_frame,
            text="📊 Processing Status & Output",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color=COLORS["text_dark"]
        )
        output_title.pack(pady=(15, 10))
        
        # Progress bar
        self.progress_var = ctk.DoubleVar()
        self.progress_bar = ctk.CTkProgressBar(
            output_frame,
            variable=self.progress_var,
            width=400,
            height=20,
            progress_color=COLORS["primary_orange"]
        )
        self.progress_bar.pack(pady=(10, 10))
        
        # Status label
        self.status_label = ctk.CTkLabel(
            output_frame,
            text="Ready to process returns data",
            font=ctk.CTkFont(size=12),
            text_color=COLORS["text_dark"]
        )
        self.status_label.pack(pady=(0, 10))
        
        # Log text area
        self.log_text = ctk.CTkTextbox(
            output_frame,
            width=600,
            height=150,
            font=ctk.CTkFont(size=11),
            fg_color=COLORS["white"],
            text_color=COLORS["text_dark"]
        )
        self.log_text.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        # Output buttons frame
        buttons_frame = ctk.CTkFrame(output_frame, fg_color="transparent")
        buttons_frame.pack(pady=(0, 15))
        
        # Open output folder button
        self.open_output_btn = ctk.CTkButton(
            buttons_frame,
            text="📁 Open Output Folder",
            command=self.open_output_folder,
            width=160,
            height=35,
            font=ctk.CTkFont(size=12),
            fg_color=COLORS["accent_blue"],
            hover_color="#1976D2",
            state="disabled"
        )
        self.open_output_btn.pack(side="left", padx=(0, 10))
        
        # Open report button  
        self.open_report_btn = ctk.CTkButton(
            buttons_frame,
            text="📋 Open Report",
            command=self.open_report,
            width=130,
            height=35,
            font=ctk.CTkFont(size=12),
            fg_color=COLORS["success_green"],
            hover_color="#388E3C",
            state="disabled"
        )
        self.open_report_btn.pack(side="left")
        
    def create_footer(self, parent):
        """Create the footer section"""
        footer_frame = ctk.CTkFrame(parent, fg_color="transparent", height=40)
        footer_frame.pack(fill="x", pady=(15, 0))
        footer_frame.pack_propagate(False)
        
        # Back to main button
        back_btn = ctk.CTkButton(
            footer_frame,
            text="← Back to Main",
            command=self.back_to_main,
            width=120,
            height=30,
            font=ctk.CTkFont(size=11),
            fg_color=COLORS["dark_gray"],
            hover_color="#555555"
        )
        back_btn.pack(side="left", pady=5)
        
        # Version info
        version_label = ctk.CTkLabel(
            footer_frame,
            text="Returns Module v1.0",
            font=ctk.CTkFont(size=10),
            text_color=COLORS["dark_gray"]
        )
        version_label.pack(side="right", pady=8)
        
    def update_file_status(self):
        """Update file status checkboxes"""
        self.log_message("Refreshing file status...")
        
        file_paths = {
            "Returns Meesho.csv": INPUT_DIR / "Returns Meesho.csv",
            "Returns Flipkart KC.csv": INPUT_DIR / "Returns Flipkart KC.csv", 
            "Returns Flipkart LL.csv": INPUT_DIR / "Returns Flipkart LL.csv",
            "Returns SellerFlex.csv": INPUT_DIR / "Returns SellerFlex.csv"
        }
        
        missing_files = []
        for filename, checkbox in self.file_checkboxes.items():
            file_exists = file_paths[filename].exists()
            checkbox.select() if file_exists else checkbox.deselect()
            
            # Update checkbox color based on file existence
            if file_exists:
                checkbox.configure(
                    fg_color=COLORS["success_green"],
                    text_color=COLORS["text_dark"]
                )
            else:
                checkbox.configure(
                    fg_color=COLORS["error_red"],
                    text_color=COLORS["text_dark"]
                )
                missing_files.append(filename)
        
        # Update process button state
        all_files_present = len(missing_files) == 0
        self.process_btn.configure(state="normal" if all_files_present else "disabled")
        
        if missing_files:
            self.log_message(f"Missing files: {', '.join(missing_files)}")
        else:
            self.log_message("✅ All required files found!")
            
    def log_message(self, message):
        """Add message to log with timestamp"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_entry = f"[{timestamp}] {message}\n"
        self.log_text.insert("end", log_entry)
        self.log_text.see("end")
        self.root.update()
        
    def start_processing(self):
        """Start the returns processing in a separate thread"""
        self.process_btn.configure(state="disabled", text="Processing...")
        self.progress_var.set(0)
        self.status_label.configure(text="Starting returns processing...")
        
        # Start processing in separate thread
        processing_thread = threading.Thread(target=self.process_returns_data)
        processing_thread.daemon = True
        processing_thread.start()
        
    def process_returns_data(self):
        """Process the returns data (main processing logic)"""
        try:
            self.log_message("🚀 Starting returns reconciliation process...")
            
            # Create output directory
            OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
            output_excel_path = OUTPUT_DIR / "Returns Reconcile Report.xlsx"
            
            self.progress_var.set(0.1)
            self.status_label.configure(text="Initializing processing...")
            
            all_data = []
            
            # Process Meesho CSV
            self.progress_var.set(0.2)
            self.status_label.configure(text="Processing Meesho data...")
            self.log_message("📄 Processing Meesho returns data...")
            
            try:
                meesho_file = INPUT_DIR / "Returns Meesho.csv"
                df_meesho = pd.read_csv(meesho_file, skiprows=7)
                selected_columns_meesho = {
                    "AWB Number": "Return TID",
                    "Type of Return": "Return Type",
                    "SKU": "SKU",
                    "Qty": "Units",
                    "Courier Partner": "Courier Partner",
                    "Order Number": "OID",
                    "Return Reason": "Cx Subject",
                    "Detailed Return Reason": "Cx Comment"
                }
                df_meesho = df_meesho[list(selected_columns_meesho.keys())].rename(columns=selected_columns_meesho)
                df_meesho["Sales Channel"] = "Meesho"
                df_meesho["Forward TID"] = ""
                all_data.append(df_meesho)
                self.log_message(f"✅ Meesho: {len(df_meesho)} records processed")
            except Exception as e:
                self.log_message(f"❌ Error processing Meesho file: {e}")
                
            # Process Flipkart KC CSV
            self.progress_var.set(0.4)
            self.status_label.configure(text="Processing Flipkart KC data...")
            self.log_message("📄 Processing Flipkart KC returns data...")
            
            try:
                flipkart_file = INPUT_DIR / "Returns Flipkart KC.csv"
                df_flipkart = pd.read_csv(flipkart_file)
                selected_columns_flipkart = {
                    "Tracking ID": "Return TID",
                    "Return Type": "Return Type",
                    "SKU": "SKU",
                    "Quantity": "Units",
                    "Order ID": "OID",
                    "Return Status": "Status of Return at the time of Capture",
                    "Return Sub-reason": "Cx Comment"
                }
                df_flipkart = df_flipkart[list(selected_columns_flipkart.keys())].rename(columns=selected_columns_flipkart)
                df_flipkart["Courier Partner"] = "Ekart"
                df_flipkart["Sales Channel"] = "Flipkart KC"
                df_flipkart["Forward TID"] = ""
                df_flipkart["Return Type"] = df_flipkart.get("Return Type", df_flipkart.get("Return Type (Column W)", ""))
                df_flipkart["Cx Subject"] = df_flipkart["Return Type"]
                all_data.append(df_flipkart)
                self.log_message(f"✅ Flipkart KC: {len(df_flipkart)} records processed")
            except Exception as e:
                self.log_message(f"❌ Error processing Flipkart KC file: {e}")
                
            # Process Flipkart LL CSV
            self.progress_var.set(0.6)
            self.status_label.configure(text="Processing Flipkart LL data...")
            self.log_message("📄 Processing Flipkart LL returns data...")
            
            try:
                flipkartll_file = INPUT_DIR / "Returns Flipkart LL.csv"
                df_flipkartll = pd.read_csv(flipkartll_file)
                selected_columns_flipkart_ll = {
                    "Tracking ID": "Return TID",
                    "Return Type": "Return Type",
                    "SKU": "SKU",
                    "Quantity": "Units",
                    "Order ID": "OID",
                    "Return Status": "Status of Return at the time of Capture",
                    "Return Sub-reason": "Cx Comment"
                }
                df_flipkartll = df_flipkartll[list(selected_columns_flipkart_ll.keys())].rename(columns=selected_columns_flipkart_ll)
                df_flipkartll["Courier Partner"] = "Ekart"
                df_flipkartll["Sales Channel"] = "Flipkart LL"
                df_flipkartll["Forward TID"] = ""
                df_flipkartll["Return Type"] = df_flipkartll.get("Return Type", df_flipkartll.get("Return Type (Column W)", ""))
                df_flipkartll["Cx Subject"] = df_flipkartll["Return Type"]
                all_data.append(df_flipkartll)
                self.log_message(f"✅ Flipkart LL: {len(df_flipkartll)} records processed")
            except Exception as e:
                self.log_message(f"❌ Error processing Flipkart LL file: {e}")
                
            # Process SellerFlex CSV
            self.progress_var.set(0.8)
            self.status_label.configure(text="Processing SellerFlex data...")
            self.log_message("📄 Processing SellerFlex returns data...")
            
            try:
                sellerflex_file = INPUT_DIR / "Returns SellerFlex.csv"
                df_sellerflex = pd.read_csv(sellerflex_file)
                selected_columns_sellerflex = {
                    "Reverse Leg Tracking ID": "Return TID",
                    "Return Type": "Return Type",
                    "mSKU": "SKU",
                    "Units": "Units",
                    "Customer Order ID": "OID",
                    "Forward Leg Tracking ID": "Forward TID",
                    "Return Status": "Status of Return at the time of Capture"
                }
                df_sellerflex = df_sellerflex[list(selected_columns_sellerflex.keys())].rename(columns=selected_columns_sellerflex)
                df_sellerflex["Courier Partner"] = "ATSIN"
                df_sellerflex["Sales Channel"] = "Amazon KC -flex"
                df_sellerflex["Cx Subject"] = ""
                df_sellerflex["Cx Comment"] = ""
                all_data.append(df_sellerflex)
                self.log_message(f"✅ SellerFlex: {len(df_sellerflex)} records processed")
            except Exception as e:
                self.log_message(f"❌ Error processing SellerFlex file: {e}")
                
            # Combine all data
            self.progress_var.set(0.9)
            self.status_label.configure(text="Consolidating data and generating report...")
            self.log_message("🔄 Combining all data sources...")
            
            if all_data:
                final_df = pd.concat(all_data, ignore_index=True)
                final_columns = [
                    "Return TID", "Return Type", "SKU", "Units", "Courier Partner",
                    "Sales Channel", "OID", "Forward TID", "Status of Return at the time of Capture",
                    "Cx Subject", "Cx Comment"
                ]
                final_df = final_df[final_columns]
                
                # Copy template to output location
                shutil.copy(TEMPLATE_PATH, output_excel_path)
                
                # Load workbook and write data
                wb = load_workbook(output_excel_path)
                ws = wb["Data"]
                
                # Write dataframe starting from A2
                for row_idx, row in enumerate(final_df.itertuples(index=False), start=2):
                    for col_idx, value in enumerate(row, start=1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Add current date to O7 and O8
                today_str = datetime.today().strftime('%d-%m-%Y')
                ws["O7"] = today_str
                ws["O8"] = today_str
                
                wb.save(output_excel_path)
                
                self.progress_var.set(1.0)
                self.status_label.configure(text="✅ Processing completed successfully!")
                self.log_message(f"✅ Total records processed: {len(final_df)}")
                self.log_message(f"✅ Output saved to: {output_excel_path}")
                
                # Enable output buttons
                self.open_output_btn.configure(state="normal")
                self.open_report_btn.configure(state="normal")
                
            else:
                raise Exception("No data was successfully processed from any source")
                
        except Exception as e:
            self.progress_var.set(0)
            self.status_label.configure(text="❌ Processing failed")
            self.log_message(f"❌ Error during processing: {str(e)}")
            
        finally:
            self.process_btn.configure(state="normal", text="🚀 Start Returns Processing")
            
    def open_output_folder(self):
        """Open the output folder in file explorer"""
        try:
            if sys.platform == "win32":
                os.startfile(OUTPUT_DIR)
            elif sys.platform == "darwin":
                subprocess.run(["open", str(OUTPUT_DIR)])
            else:
                subprocess.run(["xdg-open", str(OUTPUT_DIR)])
        except Exception as e:
            self.log_message(f"❌ Error opening output folder: {e}")
            
    def open_report(self):
        """Open the generated report"""
        try:
            report_path = OUTPUT_DIR / "Returns Reconcile Report.xlsx"
            if report_path.exists():
                if sys.platform == "win32":
                    os.startfile(report_path)
                elif sys.platform == "darwin":
                    subprocess.run(["open", str(report_path)])
                else:
                    subprocess.run(["xdg-open", str(report_path)])
            else:
                self.log_message("❌ Report file not found")
        except Exception as e:
            self.log_message(f"❌ Error opening report: {e}")
    
    def back_to_main(self):
        """Close current window and return to main"""
        if self.is_standalone:
            # If running standalone, just close the window
            self.root.destroy()
        else:
            # If integrated with homepage, restore original content
            if self.back_callback:
                self.back_callback()
            else:
                self.restore_original_content()
            
    def reset_files(self):
        """Delete all files in the input folder with confirmation"""
        # Create confirmation dialog
        confirm_dialog = ctk.CTkToplevel(self.root)
        confirm_dialog.title("Confirm Reset Files")
        confirm_dialog.geometry("400x200")
        confirm_dialog.configure(fg_color=COLORS["white"])
        
        # Make it modal
        confirm_dialog.transient(self.root)
        confirm_dialog.grab_set()
        
        # Center the dialog
        confirm_dialog.update_idletasks()
        x = (confirm_dialog.winfo_screenwidth() // 2) - 200
        y = (confirm_dialog.winfo_screenheight() // 2) - 100
        confirm_dialog.geometry(f"400x200+{x}+{y}")
        
        # Warning content
        ctk.CTkLabel(
            confirm_dialog,
            text="⚠️ Warning",
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color=COLORS["warning_orange"]
        ).pack(pady=(20, 10))
        
        ctk.CTkLabel(
            confirm_dialog,
            text="This will permanently delete ALL files in the\nReturnsreportfiles folder. This action cannot be undone.",
            font=ctk.CTkFont(size=12),
            text_color=COLORS["text_dark"],
            justify="center"
        ).pack(pady=(0, 20))
        
        # Buttons frame
        buttons_frame = ctk.CTkFrame(confirm_dialog, fg_color="transparent")
        buttons_frame.pack(pady=(0, 20))
        
        # Cancel button
        cancel_btn = ctk.CTkButton(
            buttons_frame,
            text="Cancel",
            command=confirm_dialog.destroy,
            width=100,
            height=35,
            fg_color=COLORS["dark_gray"],
            hover_color="#555555"
        )
        cancel_btn.pack(side="left", padx=(0, 10))
        
        # Confirm delete button
        def confirm_delete():
            confirm_dialog.destroy()
            self.perform_reset()
            
        delete_btn = ctk.CTkButton(
            buttons_frame,
            text="Delete All",
            command=confirm_delete,
            width=100,
            height=35,
            fg_color=COLORS["error_red"],
            hover_color="#D32F2F"
        )
        delete_btn.pack(side="left")
        
    def perform_reset(self):
        """Actually delete the files and update status"""
        try:
            deleted_count = 0
            
            # Check if input directory exists
            if INPUT_DIR.exists():
                # Get all files in the directory
                files_to_delete = list(INPUT_DIR.glob("*"))
                
                for file_path in files_to_delete:
                    if file_path.is_file():
                        try:
                            file_path.unlink()  # Delete the file
                            deleted_count += 1
                            self.log_message(f"🗑️ Deleted: {file_path.name}")
                        except Exception as e:
                            self.log_message(f"❌ Failed to delete {file_path.name}: {e}")
                            
                self.log_message(f"✅ Reset completed! {deleted_count} files deleted.")
            else:
                self.log_message("⚠️ Input directory does not exist.")
                
        except Exception as e:
            self.log_message(f"❌ Error during reset: {str(e)}")
            
        # Update file status after reset
        self.update_file_status()
        
    def run(self):
        """Start the Returns GUI application"""
        if self.is_standalone:
            self.root.mainloop()
        # If not standalone, the GUI is already displayed in the parent window

# Function to create and return the Returns module (for homepage integration)
def create_returns_module(parent_window=None, back_callback=None):
    """
    Factory function to create a Returns module instance
    Args:
        parent_window: The parent window (homepage window)
        back_callback: Function to call when back button is pressed
    Returns:
        ReturnsReportGUI instance
    """
    return ReturnsReportGUI(parent_window, back_callback)

def main():
    """Main function to run the Returns application in standalone mode"""
    app = ReturnsReportGUI()
    app.run()

# if __name__ == "__main__":
#     main()