import re
import os
import sys
import pandas as pd
import pdfplumber
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
import threading
import customtkinter as ctk
from tkinter import filedialog, messagebox
import tkinter as tk
import shutil
from pdf2image import convert_from_path
import pypdfium2

# Set CustomTkinter appearance and color theme
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

# Get the directory where the exe is located
if getattr(sys, 'frozen', False):
    BASE_DIR = Path(sys.executable).parent
else:
    BASE_DIR = Path(__file__).parent

selected_template_file = None

# Define relative paths
INPUT_DIR = BASE_DIR / "inputdir" / "PickupReportfiles"
TEMPLATE_DIR = BASE_DIR / "Template"
OUTPUT_DIR = BASE_DIR / "Output"
PIVOT_PNG_DIR = OUTPUT_DIR / "Pivot_PNGs"

# Create directories if they don't exist
INPUT_DIR.mkdir(parents=True, exist_ok=True)
TEMPLATE_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)
PIVOT_PNG_DIR.mkdir(exist_ok=True)

# --- Mapping (column → courier source) ---
prefix_mapping = {
    "A3": "Sellerflex",
    "D3": "Flipkart KC",
    "E3": "Flipkart LL",
    "F3": "Meesho - Xpressbees",
    "G3": "Meesho - Ecom Express",
    "H3": "Meesho - Delhivery",
    "I3": "Others"
}

# --- CSV input files with tracking & pivot config ---
csv_sources = {
    INPUT_DIR / "Sellerflex.csv": {
        "tracking_column": "Shipment Tracking ID",
        "source_name": "Sellerflex",
        "pivot_columns": ("MSKU", "Units")
    },
    INPUT_DIR / "Flipkart KC.csv": {
        "tracking_column": "Tracking ID",
        "source_name": "Flipkart KC",
        "pivot_columns": ("SKU", "Quantity")
    },
    INPUT_DIR / "Flipkart LL.csv": {
        "tracking_column": "Tracking ID",
        "source_name": "Flipkart LL",
        "pivot_columns": ("SKU", "Quantity")
    }
}

# GUI Color scheme - Orange and White
COLORS = {
    "primary_orange": "#FF6B35",
    "secondary_orange": "#FF8F65",
    "light_orange": "#FFB399",
    "white": "#FFFFFF",
    "light_gray": "#F5F5F5",
    "dark_gray": "#333333",
    "text_dark": "#2D2D2D",
    "success_green": "#4CAF50",
    "warning_orange": "#FF9800",
    "error_red": "#F44336",
    "accent_blue": "#1976D2"
}

class PickupReportModule:
    def __init__(self, parent_frame, back_callback=None):
        """
        Initialize the Pickup Report module
        Args:
            parent_frame: The parent frame to embed this module
            back_callback: Function to call when back button is pressed
        """
        self.parent_frame = parent_frame
        self.back_callback = back_callback
        self.processing_thread = None
        self.log_messages = []
        
        # Clear the parent frame
        for widget in self.parent_frame.winfo_children():
            widget.destroy()
            
        self.setup_gui()
        
        # Update file status on startup
        self.parent_frame.after(1000, self.update_file_status)
        
    def setup_gui(self):
        """Setup the main GUI layout"""
        # Main container
        main_frame = ctk.CTkFrame(self.parent_frame, fg_color=COLORS["white"], corner_radius=0)
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Header with back button
        header_frame = ctk.CTkFrame(main_frame, fg_color=COLORS["primary_orange"], corner_radius=10)
        header_frame.pack(fill="x", pady=(0, 20))
        
        # Back button
        if self.back_callback:
            back_btn = ctk.CTkButton(
                header_frame,
                text="← Back",
                command=self.back_callback,
                width=80,
                height=30,
                font=ctk.CTkFont(size=14, weight="bold"),
                fg_color=COLORS["accent_blue"],
                hover_color=COLORS["light_orange"]
            )
            back_btn.pack(side="left", padx=20, pady=15)
        
        # Title
        title_label = ctk.CTkLabel(
            header_frame, 
            text="Pickup Report Processing Tool",
            font=ctk.CTkFont(size=24, weight="bold"),
            text_color=COLORS["white"]
        )
        title_label.pack(pady=20)
        
        # Refresh button
        refresh_btn = ctk.CTkButton(
            header_frame,
            text="🔄",
            command=self.update_file_status,
            width=30,
            height=30,
            font=ctk.CTkFont(size=14),
            fg_color=COLORS["accent_blue"],
            hover_color=COLORS["light_orange"]
        )
        refresh_btn.pack(side="right", padx=20, pady=15)
        
        # Status section
        status_frame = ctk.CTkFrame(main_frame, fg_color=COLORS["light_gray"], corner_radius=10)
        status_frame.pack(fill="x", pady=(0, 20))
        
        status_container = ctk.CTkFrame(status_frame, fg_color="transparent")
        status_container.pack(fill="x", padx=20, pady=15)
        
        # Status text
        status_text_frame = ctk.CTkFrame(status_container, fg_color="transparent")
        status_text_frame.pack(fill="x", pady=(0, 10))
        
        ctk.CTkLabel(
            status_text_frame,
            text="Status:",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color=COLORS["text_dark"]
        ).pack(side="left")
        
        self.status_label = ctk.CTkLabel(
            status_text_frame,
            text="Ready to process",
            font=ctk.CTkFont(size=14),
            text_color=COLORS["success_green"]
        )
        self.status_label.pack(side="left", padx=(10, 0))
        
        # File status checkboxes
        files_label = ctk.CTkLabel(
            status_container,
            text="Input Files Status:",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color=COLORS["text_dark"]
        )
        files_label.pack(anchor="w", pady=(0, 5))
        
        # Create checkbox frame
        checkbox_frame = ctk.CTkFrame(status_container, fg_color="transparent")
        checkbox_frame.pack(fill="x", pady=(0, 10))
        
        # File checkboxes
        self.file_checkboxes = {}
        file_list = [
            ("Sellerflex.csv", "Sellerflex CSV"),
            ("Flipkart KC.csv", "Flipkart KC CSV"),
            ("Flipkart LL.csv", "Flipkart LL CSV"),
            ("Manifest.pdf", "Meesho PDF")
        ]
        
        # Create checkboxes in two columns
        left_frame = ctk.CTkFrame(checkbox_frame, fg_color="transparent")
        left_frame.pack(side="left", fill="x", expand=True)
        right_frame = ctk.CTkFrame(checkbox_frame, fg_color="transparent")
        right_frame.pack(side="left", fill="x", expand=True)
        
        for i, (filename, display_name) in enumerate(file_list):
            parent_frame = left_frame if i < 2 else right_frame
            
            checkbox = ctk.CTkCheckBox(
                parent_frame,
                text=display_name,
                font=ctk.CTkFont(size=12),
                text_color=COLORS["text_dark"],
                fg_color=COLORS["success_green"],
                hover_color=COLORS["light_orange"],
                checkmark_color=COLORS["white"],
                state="disabled"
            )
            checkbox.pack(anchor="w", pady=2)
            self.file_checkboxes[filename] = checkbox
        
        # Progress bar (smaller, below checkboxes)
        progress_label = ctk.CTkLabel(
            status_container,
            text="Processing Progress:",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color=COLORS["text_dark"]
        )
        progress_label.pack(anchor="w", pady=(10, 2))
        
        self.progress_bar = ctk.CTkProgressBar(
            status_container,
            height=15,
            progress_color=COLORS["secondary_orange"],
            fg_color=COLORS["white"]
        )
        self.progress_bar.pack(fill="x", pady=(0, 5))
        self.progress_bar.set(0)
        
        # Control buttons
        button_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        button_frame.pack(fill="x", pady=(0, 20))
        
        # First row of buttons
        button_row1 = ctk.CTkFrame(button_frame, fg_color="transparent")
        button_row1.pack(fill="x", pady=(0, 10))
        
        self.process_btn = ctk.CTkButton(
            button_row1,
            text="Start Processing",
            command=self.start_processing,
            width=180,
            height=40,
            font=ctk.CTkFont(size=14, weight="bold"),
            fg_color=COLORS["primary_orange"],
            hover_color=COLORS["secondary_orange"]
        )
        self.process_btn.pack(side="left", padx=(0, 10))
        
        self.clear_log_btn = ctk.CTkButton(
            button_row1,
            text="Clear Log",
            command=self.clear_log,
            width=120,
            height=40,
            font=ctk.CTkFont(size=14),
            fg_color=COLORS["light_orange"],
            hover_color=COLORS["secondary_orange"],
            text_color=COLORS["text_dark"]
        )
        self.clear_log_btn.pack(side="left", padx=(0, 10))
        
        self.output_btn = ctk.CTkButton(
            button_row1,
            text="Open Output Folder",
            command=self.open_output_folder,
            width=180,
            height=40,
            font=ctk.CTkFont(size=14),
            fg_color=COLORS["secondary_orange"],
            hover_color=COLORS["primary_orange"]
        )
        self.output_btn.pack(side="left", padx=(0, 10))
        
        # Second row of buttons
        button_row2 = ctk.CTkFrame(button_frame, fg_color="transparent")
        button_row2.pack(fill="x")
        
        self.reset_files_btn = ctk.CTkButton(
            button_row2,
            text="Reset Files",
            command=self.reset_files,
            width=120,
            height=40,
            font=ctk.CTkFont(size=14),
            fg_color=COLORS["light_orange"],
            hover_color=COLORS["secondary_orange"],
            text_color=COLORS["text_dark"]
        )
        self.reset_files_btn.pack(side="left", padx=(0, 10))

        self.structure_btn = ctk.CTkButton(
            button_row2,
            text="Show Folder Structure",
            command=self.show_folder_structure,
            width=200,
            height=40,
            font=ctk.CTkFont(size=14),
            fg_color=COLORS["light_orange"],
            hover_color=COLORS["secondary_orange"],
            text_color=COLORS["text_dark"]
        )
        self.structure_btn.pack(side="left", padx=(0, 10))
        
        self.template_btn = ctk.CTkButton(
            button_row2,
            text="Update Pickup Report",
            command=self.select_template,
            width=200,
            height=40,
            font=ctk.CTkFont(size=14),
            fg_color=COLORS["warning_orange"],
            hover_color=COLORS["secondary_orange"]
        )
        self.template_btn.pack(side="left")
        
        # Log section
        log_frame = ctk.CTkFrame(main_frame, fg_color=COLORS["light_gray"], corner_radius=10)
        log_frame.pack(fill="both", expand=True)
        
        log_label = ctk.CTkLabel(
            log_frame,
            text="Processing Log:",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color=COLORS["text_dark"]
        )
        log_label.pack(anchor="w", padx=20, pady=(15, 5))
        
        # Log text area with scrollbar
        log_container = ctk.CTkFrame(log_frame, fg_color="transparent")
        log_container.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        self.log_text = ctk.CTkTextbox(
            log_container,
            font=ctk.CTkFont(size=12, family="Consolas"),
            fg_color=COLORS["white"],
            text_color=COLORS["text_dark"],
            corner_radius=5
        )
        self.log_text.pack(fill="both", expand=True)
        self.log_text.insert("0.0", "Ready to start processing...\nClick 'Start Processing' to begin.")
        
    def log_message(self, message, level="INFO"):
        """Add message to log with timestamp"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        formatted_msg = f"[{timestamp}] {level}: {message}"
        self.log_messages.append(formatted_msg)
        print(formatted_msg)  # Also print to console
        
        # Update GUI
        self.parent_frame.after(0, self._update_log_display)
        
    def _update_log_display(self):
        """Update the log display in the GUI"""
        # Keep only last 100 messages to prevent memory issues
        if len(self.log_messages) > 100:
            self.log_messages.pop(0)
        
        self.log_text.delete("0.0", "end")
        self.log_text.insert("0.0", "\n".join(self.log_messages))
        self.log_text.see("end")
        
    def update_progress(self, value, step_description):
        """Update progress bar and current step"""
        def _update():
            self.progress_bar.set(value)
            self.status_label.configure(text=step_description)
            
        self.parent_frame.after(0, _update)
        
    def start_processing(self):
        """Start processing in a separate thread"""
        self.process_btn.configure(state="disabled")
        self.clear_log_btn.configure(state="disabled")
        
        # Clear previous log messages
        self.log_messages.clear()
        self.log_text.delete("0.0", "end")
        
        # Start processing in separate thread
        self.processing_thread = threading.Thread(target=self.process_files, daemon=True)
        self.processing_thread.start()
        
        # Start checking for completion
        self.parent_frame.after(100, self.check_processing_complete)
        
    def check_processing_complete(self):
        """Check if processing is complete and re-enable buttons"""
        if self.processing_thread and not self.processing_thread.is_alive():
            self.process_btn.configure(state="normal")
            self.clear_log_btn.configure(state="normal")
        else:
            self.parent_frame.after(100, self.check_processing_complete)
            
    def update_file_status(self):
        """Update file status checkboxes"""
        file_paths = {
            "Sellerflex.csv": INPUT_DIR / "Sellerflex.csv",
            "Flipkart KC.csv": INPUT_DIR / "Flipkart KC.csv", 
            "Flipkart LL.csv": INPUT_DIR / "Flipkart LL.csv",
            "Manifest.pdf": INPUT_DIR / "Manifest.pdf"
        }
        
        for filename, checkbox in self.file_checkboxes.items():
            file_exists = file_paths[filename].exists()
            checkbox.select() if file_exists else checkbox.deselect()
            # Update checkbox color based on file existence
            if file_exists:
                checkbox.configure(
                    fg_color=COLORS["success_green"],
                    text_color=COLORS["text_dark"]
                )
            else:
                checkbox.configure(
                    fg_color=COLORS["error_red"],
                    text_color=COLORS["text_dark"]
                )
    
    def clear_log(self):
        """Clear the log messages"""
        self.log_messages.clear()
        self.log_text.delete("0.0", "end")
        self.log_text.insert("0.0", "Log cleared. Ready for new processing...")

    def open_output_folder(self):
        """Open the output folder in file explorer"""
        try:
            if sys.platform == "win32":
                os.startfile(OUTPUT_DIR)
            elif sys.platform == "darwin":
                os.system(f"open '{OUTPUT_DIR}'")
            else:
                os.system(f"xdg-open '{OUTPUT_DIR}'")
        except Exception as e:
            self.log_message(f"Error opening folder: {e}", "ERROR")
            
    def show_folder_structure(self):
        """Show required folder structure in a dialog"""
        structure = f"""Required folder structure:

{BASE_DIR}/
├── inputdir/
│   └── PickupReportfiles/
│       ├── Sellerflex.csv
│       ├── Flipkart KC.csv
│       ├── Flipkart LL.csv
│       └── Manifest.pdf
├── Template/
│   └── Pickup Report.xlsx
├── Output/ (created automatically)
│   └── Pivot_PNGs/ (created automatically)
└── PickupReportProcessor.exe

Base Directory: {BASE_DIR}"""

        # Create a dialog window
        dialog = ctk.CTkToplevel(self.parent_frame)
        dialog.title("Folder Structure")
        dialog.geometry("650x450")
        dialog.configure(fg_color=COLORS["white"])
        
        # Make it modal
        dialog.transient(self.parent_frame)
        dialog.grab_set()
        
        # Center the dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (650 // 2)
        y = (dialog.winfo_screenheight() // 2) - (450 // 2)
        dialog.geometry(f"650x450+{x}+{y}")
        
        # Content
        text_widget = ctk.CTkTextbox(
            dialog,
            font=ctk.CTkFont(size=12, family="Consolas"),
            fg_color=COLORS["light_gray"],
            text_color=COLORS["text_dark"]
        )
        text_widget.pack(fill="both", expand=True, padx=20, pady=20)
        text_widget.insert("0.0", structure)
        text_widget.configure(state="disabled")
        
        # Close button
        close_btn = ctk.CTkButton(
            dialog,
            text="Close",
            command=dialog.destroy,
            width=100,
            height=35,
            font=ctk.CTkFont(size=14),
            fg_color=COLORS["primary_orange"],
            hover_color=COLORS["secondary_orange"]
        )
        close_btn.pack(pady=(0, 20))

    def reset_files(self):
        """Delete all files in the input directory"""
        deleted_files = 0
        for item in INPUT_DIR.iterdir():
            try:
                if item.is_file():
                    item.unlink()
                    deleted_files += 1
                elif item.is_dir():
                    shutil.rmtree(item)
                    deleted_files += 1
            except Exception as e:
                self.log_message(f"Error deleting {item.name}: {e}", "ERROR")

        self.log_message(f"Reset complete. {deleted_files} file(s)/folder(s) deleted from input directory.")
        self.update_file_status()

    def select_template(self):
        """Select template file"""
        global selected_template_file
        
        file_path = filedialog.askopenfilename(
            title="Select Pickup Report Template",
            initialdir=TEMPLATE_DIR,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if file_path:
            today_str = datetime.today().strftime("Pickup_Report_%d-%m-%Y.xlsx")
            
            if Path(file_path).name != today_str:
                messagebox.showerror(
                    "Invalid File",
                    f"Error: You have selected the wrong file.\nThe file name must be: {today_str}"
                )
                return
                
            selected_template_file = file_path
            self.log_message(f"Valid template selected: {selected_template_file}")
            messagebox.showinfo("Success", f"Template file selected successfully:\n{Path(file_path).name}")
            threading.Thread(target=self.process_files, daemon=True).start()


    # Core processing functions
    def extract_from_csv(self, path, tracking_col):
        try:
            df = pd.read_csv(path, dtype=str)
            if tracking_col not in df.columns:
                self.log_message(f"Warning: Column '{tracking_col}' not found in {path}", "WARN")
                return [], df
            tracking_ids = df[tracking_col].dropna().astype(str).tolist()
            self.log_message(f"Extracted {len(tracking_ids)} tracking IDs from {path.name}")
            return tracking_ids, df
        except FileNotFoundError:
            self.log_message(f"File not found: {path}", "ERROR")
            return [], pd.DataFrame()
        except Exception as e:
            self.log_message(f"Error reading {path}: {e}", "ERROR")
            return [], pd.DataFrame()

    def extract_data_from_pdf(self, pdf_path):
        if not pdf_path.exists():
            self.log_message(f"PDF file not found: {pdf_path}", "WARN")
            return {}
        
        known_couriers = ["Delhivery", "Ecom Express", "Xpressbees"]
        courier_data = {}

        try:
            with pdfplumber.open(pdf_path) as reader:
                total_pages = len(reader.pages[1:])
                for idx, page in enumerate(reader.pages[1:]):
                    tables = page.extract_tables()
                    page_text = page.extract_text() or ""
                    courier_name = "Unknown"

                    if "Courier :" in page_text:
                        courier_name = page_text.split("Courier :")[1].split("\n")[0].strip()

                    if courier_name not in known_couriers:
                        courier_name = "Others"

                    if courier_name not in courier_data:
                        courier_data[courier_name] = []

                    if tables:
                        for table in tables:
                            awb_col = 2
                            for row in table[1:]:
                                if len(row) > awb_col:
                                    awb = str(row[awb_col]).strip()
                                    if awb:
                                        courier_data[courier_name].append(awb)
                    
                    # Update progress for PDF processing
                    pdf_progress = 0.6 + (0.1 * (idx + 1) / total_pages)
                    self.update_progress(pdf_progress, f"Processing PDF page {idx + 1}/{total_pages}")
                    
            self.log_message(f"Extracted AWB data from PDF: {dict((k, len(v)) for k, v in courier_data.items())}")
        except Exception as e:
            self.log_message(f"Error processing PDF {pdf_path}: {e}", "ERROR")

        return courier_data

    def create_pivot_image(self, df, pivot_columns, title, filename):
        try:
            if df.empty:
                self.log_message(f"Empty DataFrame for {title}, skipping pivot creation.", "WARN")
                return
                
            sku_col, qty_col = pivot_columns
            if sku_col not in df.columns or qty_col not in df.columns:
                self.log_message(f"Required columns '{sku_col}' and '{qty_col}' not found for {title}.", "WARN")
                return

            df[qty_col] = pd.to_numeric(df[qty_col], errors='coerce').fillna(0)
            pivot = df.groupby(sku_col, as_index=False)[qty_col].sum()
            pivot.columns = ['SKU', 'Count']

            row_count = len(pivot)
            fig_height = max(4, 0.35 * (row_count + 2))
            fig, ax = plt.subplots(figsize=(8, fig_height))
            ax.axis('off')

            full_data = [[title, '']] + [pivot.columns.tolist()] + pivot.values.tolist()

            table = ax.table(
                cellText=full_data,
                loc='center',
                cellLoc='center',
                colWidths=[0.8, 0.2],
            )

            table.auto_set_font_size(False)
            table.set_fontsize(12)

            for (row, col), cell in table.get_celld().items():
                cell.set_linewidth(0.4)
                if row == 0:
                    cell.set_text_props(weight='bold', fontsize=14)
                    cell.set_fontsize(14)
                    cell.set_facecolor('#FFFFFF')
                    cell.set_height(0.5)
                    cell.visible_edges = 'open'
                elif row == 1:
                    cell.set_text_props(weight='bold')

            fig.tight_layout(pad=0)
            plt.savefig(filename, bbox_inches='tight', dpi=300, pad_inches=0.02)
            plt.close()
            self.log_message(f"Pivot table saved: {filename.name}")

        except Exception as e:
            self.log_message(f"Error creating pivot table for {title}: {e}", "ERROR")

    # def save_first_page_of_pdf_as_png(self, pdf_path, output_path):
    #     try:
    #         if not pdf_path.exists():
    #             self.log_message(f"PDF file not found: {pdf_path}", "WARN")
    #             return

    #         images = convert_from_path(str(pdf_path), first_page=1, last_page=1, dpi=150)
    #         if images:
    #             images[0].save(output_path, 'PNG')
    #             self.log_message(f"Saved first page as PNG: {output_path.name}")
    #         else:
    #             self.log_message("No pages found in the PDF", "WARN")
    #     except Exception as e:
    #         self.log_message(f"Error saving PDF page as PNG: {e}", "ERROR")
    

    def save_first_page_of_pdf_as_png(self, pdf_path, output_path):
        try:
            if not pdf_path.exists():
                self.log_message(f"PDF file not found: {pdf_path}", "WARN")
                return

            pdf = pypdfium2.PdfDocument(str(pdf_path))
            page = pdf[0]
            pil_image = page.render(scale=2).to_pil()
            pil_image.save(output_path, format="PNG")
            self.log_message(f"Saved first page as PNG: {output_path.name}")
        except Exception as e:
            self.log_message(f"Error saving PDF page as PNG: {e}", "ERROR")

    def get_top_left_if_merged(self, ws, cell_coord):
        for merged_range in ws.merged_cells.ranges:
            if cell_coord in merged_range:
                return merged_range.bounds[:2]
        return None

    def safe_write(self, ws, col, row, value):
        try:
            cell_coord = f"{col}{row}"
            merged_top_left = self.get_top_left_if_merged(ws, cell_coord)
            if merged_top_left:
                top_col, top_row = merged_top_left
                col_letter = get_column_letter(top_col)
                ws[f"{col_letter}{top_row}"] = value
            else:
                ws[cell_coord] = value
        except Exception as e:
            self.log_message(f"Error writing to cell {col}{row}: {e}", "ERROR")

    def check_required_files(self):
        """Check if all required files exist"""
        TEMPLATE_FILE = Path(selected_template_file) if selected_template_file else TEMPLATE_DIR / "Pickup Report.xlsx"
        meesho_pdf = INPUT_DIR / "Manifest.pdf"
        
        missing_files = []
        
        if not TEMPLATE_FILE.exists():
            missing_files.append(f"Template: {TEMPLATE_FILE}")
        
        for csv_path in csv_sources.keys():
            if not csv_path.exists():
                missing_files.append(f"CSV: {csv_path}")
        
        if not meesho_pdf.exists():
            missing_files.append(f"PDF: {meesho_pdf}")
        
        return missing_files, TEMPLATE_FILE, meesho_pdf

    def process_files(self):
        """Main processing function that runs in a separate thread"""
        try:
            self.log_message("Starting Pickup Report Processing...")
            self.update_progress(0.05, "Checking required files...")

            missing_files, TEMPLATE_FILE, meesho_pdf = self.check_required_files()

            if not TEMPLATE_FILE or not TEMPLATE_FILE.exists():
                self.log_message("Template file is missing or invalid.", "ERROR")
                self.update_progress(0.0, "Missing template file!")
                return

            # We allow other files to be missing; only warn
            if missing_files:
                self.log_message("Some input files are missing; proceeding with available files.", "WARN")
                for file in missing_files:
                    self.log_message(f"  - Missing: {file}", "WARN")

            self.log_message("Proceeding with available files...")
            self.update_progress(0.1, "Loading template...")

            # Modified: Determine output file based on whether template was manually selected
            if selected_template_file:
                # If user manually selected template, save to that template file
                OUTPUT_FILE = Path(selected_template_file)
                self.log_message(f"Using manually selected template: {OUTPUT_FILE.name}")
            else:
                # Default behavior: save to output directory
                OUTPUT_FILE = OUTPUT_DIR / f"Pickup_Report_{datetime.today().strftime('%d-%m-%Y')}.xlsx"

            START_ROW = 3

            wb = load_workbook(TEMPLATE_FILE)
            if "Entry tracking ID" not in wb.sheetnames:
                raise ValueError("Sheet 'Entry tracking ID' not found in template.")

            ws = wb["Entry tracking ID"]
            ws["K1"] = datetime.today().strftime('%d-%m-%Y')

            col_for = {v: re.match(r"([A-Z]+)", k).group(1) for k, v in prefix_mapping.items()}

            # Process CSVs that are available
            total_sources = sum(1 for path in csv_sources if path.exists())
            current_idx = 0

            for path, config in csv_sources.items():
                if not path.exists():
                    self.log_message(f"Skipping missing source: {path.name}", "WARN")
                    continue

                current_idx += 1
                tracking_col = config["tracking_column"]
                source_name = config["source_name"]
                pivot_columns = config["pivot_columns"]

                progress = 0.1 + (0.4 * current_idx / total_sources)
                self.update_progress(progress, f"Processing {source_name}...")

                self.log_message(f"Processing {source_name} from {path.name}")
                tracking_ids, df = self.extract_from_csv(path, tracking_col)

                if not tracking_ids:
                    self.log_message(f"No tracking IDs found for {source_name}", "WARN")
                    continue

                # Remove duplicates
                tracking_ids = list(dict.fromkeys(tracking_ids))

                col = col_for.get(source_name)
                if not col:
                    self.log_message(f"No column mapping found for {source_name}", "WARN")
                    continue

                for i, tid in enumerate(tracking_ids, START_ROW):
                    val = tid if not str(tid).isdigit() else int(tid)
                    self.safe_write(ws, col, i, val)

                img_path = PIVOT_PNG_DIR / f"{source_name.replace(' ', '_')}_pivot.png"
                self.create_pivot_image(df, pivot_columns, f"{source_name} Pivot Table", img_path)

            # Meesho PDF Processing
            if meesho_pdf and meesho_pdf.exists():
                self.update_progress(0.6, "Processing Meesho PDF...")
                self.log_message(f"Processing Meesho PDF: {meesho_pdf.name}")
                courier_data = self.extract_data_from_pdf(meesho_pdf)

                for courier, awbs in courier_data.items():
                    column_key = f"Meesho - {courier}" if courier != "Others" else "Others"
                    if column_key not in col_for:
                        self.log_message(f"Skipping unrecognized courier: {courier}", "WARN")
                        continue
                    col = col_for[column_key]

                    # Remove duplicates
                    awbs = list(dict.fromkeys(awbs))

                    for i, tid in enumerate(awbs, START_ROW):
                        self.safe_write(ws, col, i, tid)

                self.update_progress(0.8, "Creating Meesho pivot image...")
                meesho_pivot_image_path = PIVOT_PNG_DIR / "Meesho_Pivot_Page.png"
                self.save_first_page_of_pdf_as_png(meesho_pdf, meesho_pivot_image_path)
            else:
                self.log_message("Meesho PDF not found. Skipping PDF processing.", "WARN")

            # Final steps
            self.update_progress(0.9, "Saving Excel file...")
            wb.save(OUTPUT_FILE)

            self.update_progress(1.0, "Processing completed successfully!")
            self.log_message(f"Pickup report saved: {OUTPUT_FILE.name}")
            self.log_message(f"All pivot tables saved in: {PIVOT_PNG_DIR}")
            self.log_message("Processing completed successfully!", "SUCCESS")

        except Exception as e:
            self.log_message(f"Fatal error: {e}", "ERROR")
            self.update_progress(0.0, "Processing failed!")

    def run(self):
        """Start the GUI application"""
        self.root.mainloop()

def main():
    """Main function to run the application"""
    # Add matplotlib backend for headless operation
    plt.switch_backend('Agg')
    
    # Create main window
    root = ctk.CTk()
    root.title("Pickup Report Processing Tool")
    root.geometry("1200x800")
    root.configure(fg_color=COLORS["white"])
    
    # Create main frame
    main_frame = ctk.CTkFrame(root, fg_color="transparent")
    main_frame.pack(fill="both", expand=True)
    
    # Create and run the pickup report module
    app = PickupReportModule(main_frame)
    
    # Start the main event loop
    root.mainloop()

if __name__ == "__main__":
    main()